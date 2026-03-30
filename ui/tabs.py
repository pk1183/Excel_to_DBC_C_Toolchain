"""
tabs.py - Tab widgets for Excel to DBC to C Toolchain UI
"""
import os
import openpyxl
from PyQt6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QLabel, QPushButton,
    QLineEdit, QTableWidget, QTableWidgetItem, QTextEdit,
    QFileDialog, QSplitter, QHeaderView, QFrame, QComboBox,
    QMessageBox
)
from PyQt6.QtCore import Qt, QThread, pyqtSignal
from PyQt6.QtGui import QColor, QFont
from ui.dialogs.can_row_dialog import CanRowDialog


# ─────────────────────────────────────────────────────────────
# Tab 1: Excel Populating & Review
# ─────────────────────────────────────────────────────────────
class ExcelTab(QWidget):
    """Tab for browsing and previewing the Excel CAN matrix."""

    # Signal emitted (excel_path) when user confirms they want to generate DBC
    generate_dbc_requested = pyqtSignal(str)
    visualizer_update_requested = pyqtSignal(str, list, str) # message_name, signals, selected_signal_name
    visualizer_matrix_updated = pyqtSignal(dict) # {message_name: [signal_dicts]}

    def __init__(self, parent=None):
        super().__init__(parent)
        self._excel_path = ""
        self._build_ui()

    def _build_ui(self):
        root = QVBoxLayout(self)
        root.setContentsMargins(16, 16, 16, 16)
        root.setSpacing(12)

        # ── Section header ──────────────────────────────────────
        hdr = QLabel("Excel CAN Matrix")
        hdr.setObjectName("sectionTitle")
        root.addWidget(hdr)

        # ── File picker ─────────────────────────────────────────
        card = QFrame()
        card.setObjectName("panelCard")
        card_layout = QVBoxLayout(card)
        card_layout.setContentsMargins(16, 14, 16, 14)
        card_layout.setSpacing(10)

        picker_row = QHBoxLayout()

        self.path_edit = QLineEdit()
        self.path_edit.setPlaceholderText("Select Excel (.xlsx) file …")
        self.path_edit.setReadOnly(True)

        browse_btn = QPushButton("Browse …")
        browse_btn.setFixedWidth(100)
        browse_btn.clicked.connect(self._browse_file)

        picker_row.addWidget(QLabel("Input File:"))
        picker_row.addWidget(self.path_edit, 1)
        picker_row.addWidget(browse_btn)
        card_layout.addLayout(picker_row)

        # ── Sheet selector ──────────────────────────────────────
        sheet_row = QHBoxLayout()
        self.sheet_combo = QComboBox()
        self.sheet_combo.setMinimumWidth(200)
        self.sheet_combo.currentTextChanged.connect(self._load_sheet)
        sheet_row.addWidget(QLabel("Sheet:"))
        sheet_row.addWidget(self.sheet_combo)
        sheet_row.addStretch(1)

        self.row_count_label = QLabel("—")
        sheet_row.addWidget(self.row_count_label)
        sheet_row.addSpacing(20)

        self.add_entry_btn = QPushButton("Add Entry")
        self.add_entry_btn.setObjectName("PrimaryButton")
        self.add_entry_btn.setFixedWidth(100)
        self.add_entry_btn.setEnabled(False)
        self.add_entry_btn.clicked.connect(self._on_add_entry)
        sheet_row.addWidget(self.add_entry_btn)

        self.modify_entry_btn = QPushButton("Modify Entry")
        self.modify_entry_btn.setObjectName("SecondaryButton")
        self.modify_entry_btn.setFixedWidth(100)
        self.modify_entry_btn.setEnabled(False)
        self.modify_entry_btn.clicked.connect(self._on_modify_entry)
        sheet_row.addWidget(self.modify_entry_btn)

        card_layout.addLayout(sheet_row)
        root.addWidget(card)

        # ── Table preview ───────────────────────────────────────
        self.table = QTableWidget()
        self.table.setAlternatingRowColors(True)
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Interactive)
        self.table.horizontalHeader().setStretchLastSection(True)
        self.table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self.table.setSelectionMode(QTableWidget.SelectionMode.SingleSelection)
        self.table.itemSelectionChanged.connect(self._on_table_selection_changed)
        root.addWidget(self.table, 1)

    # ── Slots ───────────────────────────────────────────────────
    def _browse_file(self):
        path, _ = QFileDialog.getOpenFileName(
            self, "Open Excel File", "", "Excel Files (*.xlsx *.xls)"
        )
        if path:
            self._excel_path = path
            self.path_edit.setText(path)
            self._load_workbook(path)

    def _load_workbook(self, path: str):
        try:
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
            self.sheet_combo.blockSignals(True)
            self.sheet_combo.clear()
            self.sheet_combo.addItems(wb.sheetnames)
            self.sheet_combo.blockSignals(False)
            if wb.sheetnames:
                self._load_sheet(wb.sheetnames[0])
            self.add_entry_btn.setEnabled(True)
        except Exception as e:
            self.add_entry_btn.setEnabled(False)
            self._show_error_in_table(str(e))

    def _load_sheet(self, sheet_name: str):
        if not self._excel_path or not sheet_name:
            return
        try:
            wb = openpyxl.load_workbook(self._excel_path, read_only=True, data_only=True)
            ws = wb[sheet_name]

            all_rows = list(ws.iter_rows(values_only=True))
            if not all_rows:
                return

            headers = [str(c) if c is not None else "" for c in all_rows[0]]
            data_rows = all_rows[1:]

            self.table.setColumnCount(len(headers))
            self.table.setHorizontalHeaderLabels(headers)
            self.table.setRowCount(len(data_rows))

            for r_idx, row in enumerate(data_rows):
                for c_idx, cell in enumerate(row):
                    text = str(cell) if cell is not None else ""
                    item = QTableWidgetItem(text)
                    item.setFlags(item.flags() & ~Qt.ItemFlag.ItemIsEditable)
                    # Highlight non-empty rows with a subtle green tint
                    if any(v is not None and str(v).strip() for v in row):
                        item.setBackground(QColor("#F0FDF4"))
                    self.table.setItem(r_idx, c_idx, item)

            self.row_count_label.setText(f"{len(data_rows)} rows")
            
            # Form Global Matrix for Visualizer
            matrix_dict = {}
            col_map = {self.table.horizontalHeaderItem(c).text().strip(): c for c in range(self.table.columnCount()) if self.table.horizontalHeaderItem(c)}
            
            alias_map = {
                'Message Name': ['Message Name', 'Message', 'Name'],
                'Signal Name': ['Signal Name', 'Signal'],
                'Start Bit': ['Start Bit', 'StartBit', 'Pos', 'Position'],
                'Length': ['Length', 'Size', 'Len', 'Bits'],
                'Byte Order': ['Byte Order', 'ByteOrder', 'Endian']
            }
            def _find_col(name):
                for cand in alias_map.get(name, [name]):
                    if cand in col_map: return col_map[cand]
                return -1

            msg_col = _find_col('Message Name')
            sig_col = _find_col('Signal Name')
            sb_col, len_col, bo_col = _find_col('Start Bit'), _find_col('Length'), _find_col('Byte Order')

            if msg_col != -1:
                for r in range(self.table.rowCount()):
                    m_item = self.table.item(r, msg_col)
                    if m_item and m_item.text().strip():
                        m_name = m_item.text().strip()
                        if m_name not in matrix_dict:
                            matrix_dict[m_name] = []
                        
                        sig_data = {}
                        if sig_col != -1 and self.table.item(r, sig_col): sig_data['Signal Name'] = self.table.item(r, sig_col).text().strip()
                        if sb_col != -1 and self.table.item(r, sb_col): sig_data['Start Bit'] = self.table.item(r, sb_col).text().strip()
                        if len_col != -1 and self.table.item(r, len_col): sig_data['Length'] = self.table.item(r, len_col).text().strip()
                        if bo_col != -1 and self.table.item(r, bo_col): sig_data['Byte Order'] = self.table.item(r, bo_col).text().strip()
                        matrix_dict[m_name].append(sig_data)

            self.visualizer_matrix_updated.emit(matrix_dict)

        except Exception as e:
            self._show_error_in_table(str(e))

    def _show_error_in_table(self, msg: str):
        self.table.setColumnCount(1)
        self.table.setHorizontalHeaderLabels(["Error"])
        self.table.setRowCount(1)
        item = QTableWidgetItem(msg)
        item.setForeground(QColor("#DC2626"))
        self.table.setItem(0, 0, item)

    def _on_table_selection_changed(self):
        selected_indexes = self.table.selectedIndexes()
        has_selection = len(selected_indexes) > 0
        file_loaded = self.add_entry_btn.isEnabled()
        self.modify_entry_btn.setEnabled(file_loaded and has_selection)

        if has_selection:
            row = selected_indexes[0].row()
            
            # Map columns
            col_map = {}
            for c in range(self.table.columnCount()):
                header_item = self.table.horizontalHeaderItem(c)
                if header_item:
                    col_map[header_item.text().strip()] = c
            
            alias_map = {
                'Message Name': ['Message Name', 'Message', 'Name'],
                'Signal Name': ['Signal Name', 'Signal'],
                'Start Bit': ['Start Bit', 'StartBit', 'Pos', 'Position'],
                'Length': ['Length', 'Size', 'Len', 'Bits'],
                'Byte Order': ['Byte Order', 'ByteOrder', 'Endian']
            }

            def _find_col(logical_name):
                for candidate in alias_map.get(logical_name, [logical_name]):
                    if candidate in col_map:
                        return col_map[candidate]
                return -1

            msg_name_col = _find_col('Message Name')
            sig_name_col = _find_col('Signal Name')
            start_bit_col = _find_col('Start Bit')
            length_col = _find_col('Length')
            byte_order_col = _find_col('Byte Order')

            if msg_name_col != -1:
                item = self.table.item(row, msg_name_col)
                if item:
                    target_msg_name = item.text().strip()
                    selected_sig = ""
                    if sig_name_col != -1 and self.table.item(row, sig_name_col):
                        selected_sig = self.table.item(row, sig_name_col).text().strip()
                    
                    signals = []
                    for r in range(self.table.rowCount()):
                        m_item = self.table.item(r, msg_name_col)
                        if m_item and m_item.text().strip() == target_msg_name:
                            sig_data = {}
                            if sig_name_col != -1 and self.table.item(r, sig_name_col):
                                sig_data['Signal Name'] = self.table.item(r, sig_name_col).text().strip()
                            if start_bit_col != -1 and self.table.item(r, start_bit_col):
                                sig_data['Start Bit'] = self.table.item(r, start_bit_col).text().strip()
                            if length_col != -1 and self.table.item(r, length_col):
                                sig_data['Length'] = self.table.item(r, length_col).text().strip()
                            if byte_order_col != -1 and self.table.item(r, byte_order_col):
                                sig_data['Byte Order'] = self.table.item(r, byte_order_col).text().strip()
                            
                            signals.append(sig_data)
                    
                    self.visualizer_update_requested.emit(target_msg_name, signals, selected_sig)

    def _gather_existing_data(self):
        existing_messages = {}
        existing_signals = []
        col_map = {}
        for c in range(self.table.columnCount()):
            header_item = self.table.horizontalHeaderItem(c)
            if header_item:
                col_map[header_item.text().strip()] = c
        
        # Flexible column mapping rules
        alias_map = {
            'Message ID': ['Message ID (Hex)', 'ID (Hex)', 'Message ID', 'ID', 'CAN ID'],
            'Message Name': ['Message Name', 'Message', 'Name'],
            'ID Type': ['ID Type', 'Type', 'Frame Type'],
            'Cycle Time': ['Cycle Time (ms)', 'Cycle Time', 'Period'],
            'Signal Name': ['Signal Name', 'Signal'],
        }

        def _find_col(logical_name):
            for candidate in alias_map.get(logical_name, [logical_name]):
                if candidate in col_map:
                    return col_map[candidate]
            return -1

        msg_name_col = _find_col("Message Name")
        msg_id_col = _find_col("Message ID")
        sig_name_col = _find_col("Signal Name")
        id_type_col = _find_col("ID Type")
        cycle_col = _find_col("Cycle Time")
        
        # Populate existing data
        if msg_name_col != -1 and msg_id_col != -1 and sig_name_col != -1:
            for r in range(self.table.rowCount()):
                msg_name_item = self.table.item(r, msg_name_col)
                msg_id_item = self.table.item(r, msg_id_col)
                sig_name_item = self.table.item(r, sig_name_col)

                if msg_name_item and msg_id_item:
                    m_name = msg_name_item.text().strip()
                    m_id = msg_id_item.text().strip()
                    if m_name and m_id:
                        if m_name not in existing_messages:
                            id_type_val = self.table.item(r, id_type_col) if id_type_col != -1 else None
                            id_type_val = id_type_val.text().strip() if id_type_val else ""
                            cycle_val = self.table.item(r, cycle_col) if cycle_col != -1 else None
                            cycle_val = cycle_val.text().strip() if cycle_val else ""
                            existing_messages[m_name] = {
                                "Message ID": m_id,
                                "ID Type": id_type_val,
                                "Cycle Time": cycle_val
                            }
                if sig_name_item:
                    s_name = sig_name_item.text().strip()
                    if s_name:
                        existing_signals.append(s_name)
        
        return existing_messages, existing_signals, col_map

    def _on_add_entry(self):
        if not self._excel_path or not os.path.exists(self._excel_path):
            QMessageBox.warning(self, "No File", "Please select a valid Excel file first.")
            return

        sheet_name = self.sheet_combo.currentText()
        if not sheet_name:
            return

        existing_messages, existing_signals, _ = self._gather_existing_data()

        dialog = CanRowDialog(self, existing_messages, existing_signals)
        dialog.save_requested.connect(self._save_excel_data)
        dialog.exec()

    def _on_modify_entry(self):
        if not self._excel_path or not os.path.exists(self._excel_path):
            QMessageBox.warning(self, "No File", "Please select a valid Excel file first.")
            return

        sheet_name = self.sheet_combo.currentText()
        if not sheet_name:
            return

        row = self.table.currentRow()
        if row < 0:
            return

        existing_messages, existing_signals, col_map = self._gather_existing_data()

        # Build edit_data based on alias mapping
        alias_map = {
            'Message ID': ['Message ID (Hex)', 'ID (Hex)', 'Message ID', 'ID', 'CAN ID'],
            'Message Name': ['Message Name', 'Message', 'Name'],
            'ID Type': ['ID Type', 'Type', 'Frame Type'],
            'Cycle Time': ['Cycle Time (ms)', 'Cycle Time', 'Period'],
            'Signal Name': ['Signal Name', 'Signal'],
            'Start Bit': ['Start Bit', 'StartBit', 'Pos', 'Position'],
            'Length': ['Length', 'Size', 'Len', 'Bits'],
            'Byte Order': ['Byte Order', 'ByteOrder', 'Endian'],
            'Factor': ['Factor', 'Scale'],
            'Offset': ['Offset'],
            'Min': ['Min', 'Minimum'],
            'Max': ['Max', 'Maximum'],
            'Unit': ['Unit', 'Units'],
            'Value Descriptions': ['Value Descriptions', 'Values', 'Choices', 'Enum']
        }

        edit_data = {}
        for logical_cap, candidates in alias_map.items():
            for c in candidates:
                if c in col_map:
                    item = self.table.item(row, col_map[c])
                    edit_data[logical_cap] = item.text().strip() if item else ""
                    break

        dialog = CanRowDialog(self, existing_messages, existing_signals, edit_data=edit_data, update_row=row)
        dialog.save_requested.connect(self._save_excel_data)
        dialog.exec()

    def _save_excel_data(self, data, is_update, update_row):
        sheet_name = self.sheet_combo.currentText()
        try:
            wb = openpyxl.load_workbook(self._excel_path)
            if sheet_name not in wb.sheetnames:
                QMessageBox.critical(self, "Error", f"Sheet {sheet_name} not found.")
                return
            ws = wb[sheet_name]

            # Find columns
            header_row = None
            header_idx = -1
            for idx, row in enumerate(ws.iter_rows(values_only=True)):
                if row and any(row):
                    header_row = [str(c).strip() if c is not None else "" for c in row]
                    header_idx = idx + 1
                    break
            
            if not header_row:
                QMessageBox.critical(self, "Error", "Could not find header row in sheet.")
                return
                
            col_map = {name: i for i, name in enumerate(header_row)}

            # Full flexible column mapping
            alias_map = {
                'Message ID': ['Message ID (Hex)', 'ID (Hex)', 'Message ID', 'ID', 'CAN ID'],
                'Message Name': ['Message Name', 'Message', 'Name'],
                'ID Type': ['ID Type', 'Type', 'Frame Type'],
                'Cycle Time': ['Cycle Time (ms)', 'Cycle Time', 'Period'],
                'Signal Name': ['Signal Name', 'Signal'],
                'Start Bit': ['Start Bit', 'StartBit', 'Pos', 'Position'],
                'Length': ['Length', 'Size', 'Len', 'Bits'],
                'Byte Order': ['Byte Order', 'ByteOrder', 'Endian'],
                'Factor': ['Factor', 'Scale'],
                'Offset': ['Offset'],
                'Min': ['Min', 'Minimum'],
                'Max': ['Max', 'Maximum'],
                'Unit': ['Unit', 'Units'],
                'Multiplex Type': ['Multiplex Type', 'Mux Type', 'Multiplex'],
                'Multiplex Value': ['Multiplex Value', 'Mux Value'],
                'Value Descriptions': ['Value Descriptions', 'Values', 'Choices', 'Enum']
            }
            
            def _find_header_idx(logical_name):
                for candidate in alias_map.get(logical_name, [logical_name]):
                    if candidate in col_map:
                        return col_map[candidate]
                return -1
                
            # Create a row matching the header order
            new_row = [""] * len(header_row)
            for k, v in data.items():
                target_idx = _find_header_idx(k)
                if target_idx != -1:
                    new_row[target_idx] = v
                else:
                    # Column not present in excel sheet
                    pass
            
            if is_update and update_row >= 0:
                row_idx = header_idx + update_row + 1
                for idx, val in enumerate(new_row):
                    ws.cell(row=row_idx, column=idx + 1, value=val)
            else:
                ws.append(new_row)
            
            wb.save(self._excel_path)
            
            # Reload view
            self._load_sheet(sheet_name)
            
            if is_update:
                self.table.selectRow(update_row)
                QMessageBox.information(self, "Success", "Modified entry in Excel successfully.")
            else:
                self.table.scrollToBottom()
                QMessageBox.information(self, "Success", "Added new entry to Excel successfully.")

        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to save to Excel:\n{e}")


    def get_excel_path(self) -> str:
        return self._excel_path

    def load_from_path(self, path: str):
        """Pre-load a file (called from main window when config is read)."""
        if path and os.path.exists(path):
            self._excel_path = path
            self.path_edit.setText(path)
            self._load_workbook(path)


# ─────────────────────────────────────────────────────────────
# Tab 2: DBC Review
# ─────────────────────────────────────────────────────────────
class DbcTab(QWidget):
    """Tab for reviewing the generated .dbc file."""

    generate_code_requested = pyqtSignal()
    export_excel_requested = pyqtSignal(str) # emits dbc_path
    load_dbc_requested = pyqtSignal(str)     # emits dbc_path

    def __init__(self, parent=None):
        super().__init__(parent)
        self._dbc_path = ""
        self._build_ui()

    def _build_ui(self):
        root = QVBoxLayout(self)
        root.setContentsMargins(16, 16, 16, 16)
        root.setSpacing(12)

        hdr = QLabel("DBC File Review")
        hdr.setObjectName("sectionTitle")
        root.addWidget(hdr)

        # ── Info card ───────────────────────────────────────────
        card = QFrame()
        card.setObjectName("panelCard")
        card_layout = QHBoxLayout(card)
        card_layout.setContentsMargins(16, 12, 16, 12)

        self.dbc_path_label = QLabel("No DBC file loaded.")
        self.dbc_path_label.setWordWrap(True)
        card_layout.addWidget(QLabel("File:"))
        card_layout.addWidget(self.dbc_path_label, 1)

        button_layout = QHBoxLayout()
        button_layout.setSpacing(8)

        load_btn = QPushButton("Load DBC")
        load_btn.clicked.connect(self._on_load_clicked)
        button_layout.addWidget(load_btn)

        reload_btn = QPushButton("Reload")
        reload_btn.clicked.connect(self._reload)
        button_layout.addWidget(reload_btn)

        self.export_excel_btn = QPushButton("Export to Excel Matrix")
        self.export_excel_btn.clicked.connect(self._on_export_excel_clicked)
        button_layout.addWidget(self.export_excel_btn)

        self.gen_code_btn = QPushButton("Generate C/H Code")
        self.gen_code_btn.clicked.connect(lambda: self.generate_code_requested.emit())
        button_layout.addWidget(self.gen_code_btn)

        card_layout.addLayout(button_layout)
        root.addWidget(card)

        # ── DBC text viewer ─────────────────────────────────────
        self.text_view = QTextEdit()
        self.text_view.setReadOnly(True)
        self.text_view.setPlaceholderText("DBC content will appear here after generation …")
        root.addWidget(self.text_view, 1)

    def _on_load_clicked(self):
        path, _ = QFileDialog.getOpenFileName(self, "Load DBC File", "", "DBC Files (*.dbc)")
        if path:
            self.load_dbc(path)
            self.load_dbc_requested.emit(path)

    def _on_export_excel_clicked(self):
        if not self._dbc_path or not os.path.exists(self._dbc_path):
            QMessageBox.warning(self, "No DBC Loaded", "Please load or generate a DBC file first.")
            return
        self.export_excel_requested.emit(self._dbc_path)

    def load_dbc(self, dbc_path: str):
        self._dbc_path = dbc_path
        self._reload()

    def _reload(self):
        if not self._dbc_path or not os.path.exists(self._dbc_path):
            self.text_view.setPlaceholderText("DBC file not found. Generate DBC first.")
            return
        try:
            with open(self._dbc_path, "r", encoding="utf-8", errors="replace") as f:
                content = f.read()
            self.text_view.setPlainText(content)
            self.dbc_path_label.setText(self._dbc_path)
        except Exception as e:
            self.text_view.setPlainText(f"Error reading file:\n{e}")

    def get_dbc_path(self) -> str:
        return self._dbc_path


# ─────────────────────────────────────────────────────────────
# Tab 3: C / H Code Review
# ─────────────────────────────────────────────────────────────
class CodeTab(QWidget):
    """Tab for reviewing generated .c and .h files (split view)."""

    def __init__(self, parent=None):
        super().__init__(parent)
        self._output_dir = ""
        self._build_ui()

    def _build_ui(self):
        root = QVBoxLayout(self)
        root.setContentsMargins(16, 16, 16, 16)
        root.setSpacing(12)

        hdr = QLabel("Generated C / H Code Review")
        hdr.setObjectName("sectionTitle")
        root.addWidget(hdr)

        # ── File selectors ──────────────────────────────────────
        card = QFrame()
        card.setObjectName("panelCard")
        card_layout = QHBoxLayout(card)
        card_layout.setContentsMargins(16, 10, 16, 10)
        card_layout.setSpacing(12)

        self.c_selector = QComboBox()
        self.c_selector.setMinimumWidth(220)
        self.c_selector.currentTextChanged.connect(self._load_c_file)

        self.h_selector = QComboBox()
        self.h_selector.setMinimumWidth(220)
        self.h_selector.currentTextChanged.connect(self._load_h_file)

        card_layout.addWidget(QLabel(".c file:"))
        card_layout.addWidget(self.c_selector)
        card_layout.addSpacing(20)
        card_layout.addWidget(QLabel(".h file:"))
        card_layout.addWidget(self.h_selector)
        card_layout.addStretch(1)
        root.addWidget(card)

        # ── Split viewer ────────────────────────────────────────
        splitter = QSplitter(Qt.Orientation.Horizontal)

        left_panel = QWidget()
        left_layout = QVBoxLayout(left_panel)
        left_layout.setContentsMargins(0, 0, 0, 0)
        left_layout.setSpacing(4)
        left_label = QLabel(".c — Implementation")
        left_label.setStyleSheet("color: #64748B; font-size: 9pt;")
        self.c_view = QTextEdit()
        self.c_view.setReadOnly(True)
        self.c_view.setPlaceholderText("C source file will appear here …")
        left_layout.addWidget(left_label)
        left_layout.addWidget(self.c_view)

        right_panel = QWidget()
        right_layout = QVBoxLayout(right_panel)
        right_layout.setContentsMargins(0, 0, 0, 0)
        right_layout.setSpacing(4)
        right_label = QLabel(".h — Header")
        right_label.setStyleSheet("color: #64748B; font-size: 9pt;")
        self.h_view = QTextEdit()
        self.h_view.setReadOnly(True)
        self.h_view.setPlaceholderText("Header file will appear here …")
        right_layout.addWidget(right_label)
        right_layout.addWidget(self.h_view)

        splitter.addWidget(left_panel)
        splitter.addWidget(right_panel)
        splitter.setSizes([500, 500])
        root.addWidget(splitter, 1)

    def load_output_dir(self, output_dir: str):
        self._output_dir = output_dir
        self.c_selector.blockSignals(True)
        self.h_selector.blockSignals(True)
        self.c_selector.clear()
        self.h_selector.clear()

        if not os.path.isdir(output_dir):
            self.c_selector.blockSignals(False)
            self.h_selector.blockSignals(False)
            return

        c_files = sorted(f for f in os.listdir(output_dir) if f.endswith(".c"))
        h_files = sorted(f for f in os.listdir(output_dir) if f.endswith(".h"))

        self.c_selector.addItems(c_files)
        self.h_selector.addItems(h_files)

        self.c_selector.blockSignals(False)
        self.h_selector.blockSignals(False)

        if c_files:
            self._load_c_file(c_files[0])
        if h_files:
            self._load_h_file(h_files[0])

    def _load_c_file(self, filename: str):
        self._load_into(self.c_view, filename)

    def _load_h_file(self, filename: str):
        self._load_into(self.h_view, filename)

    def _load_into(self, text_edit: QTextEdit, filename: str):
        if not filename or not self._output_dir:
            return
        path = os.path.join(self._output_dir, filename)
        try:
            with open(path, "r", encoding="utf-8", errors="replace") as f:
                text_edit.setPlainText(f.read())
        except Exception as e:
            text_edit.setPlainText(f"Error reading {filename}:\n{e}")
