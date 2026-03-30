import cantools
import openpyxl
import sys
import os

def dbc_to_excel(dbc_path, output_path):
    """
    Parses a DBC file and meticulously rebuilds a neatly formatted Excel matrix 
    matching the unified CAN toolchain specifications.
    """
    if not os.path.exists(dbc_path):
        print(f"[ERROR] DBC file not found at {dbc_path}")
        sys.exit(1)

    print(f"[INFO] Loading DBC: {dbc_path}")
    try:
        db = cantools.database.load_file(dbc_path)
    except Exception as e:
        print(f"[ERROR] Error loading DBC: {e}")
        sys.exit(1)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "CAN Matrix"

    # Define toolchain compliant headers
    headers = [
        "Message Name", "ID (Hex)", "ID Type", "Cycle Time", "Signal Name",
        "Start Bit", "Length", "Byte Order", "Factor", "Offset", "Min", "Max",
        "Unit", "Multiplex Type", "Multiplex Value", "Value Descriptions"
    ]
    ws.append(headers)

    message_count = 0
    signal_count = 0

    print(f"[INFO] Iterating mapping logic across objects...")
    for msg in db.messages:
        message_count += 1
        msg_name = msg.name
        msg_id_hex = f"0x{msg.frame_id:X}"
        id_type = "Extended" if msg.is_extended_frame else "Standard"
        cycle_time = msg.cycle_time if msg.cycle_time else 0

        if not msg.signals:
            # Append empty message
            ws.append([msg_name, msg_id_hex, id_type, cycle_time, "", "", "", "", "", "", "", "", "", "", "", ""])

        for sig in msg.signals:
            signal_count += 1
            sig_name = sig.name
            start_bit = sig.start
            length = sig.length
            byte_order = "Big Endian" if sig.byte_order == "big_endian" else "Little Endian"
            factor = float(sig.scale) if sig.scale is not None else 1.0
            offset = float(sig.offset) if sig.offset is not None else 0.0
            minimum = float(sig.minimum) if sig.minimum is not None else 0.0
            maximum = float(sig.maximum) if sig.maximum is not None else 0.0
            unit = sig.unit if sig.unit else ""

            mux_type = ""
            mux_val = ""
            if sig.is_multiplexer:
                mux_type = "M"
            elif sig.multiplexer_ids:
                mux_type = "m"
                mux_val = sig.multiplexer_ids[0]

            val_desc = ""
            if sig.choices:
                val_desc = ", ".join([f"{k}:{v}" for k, v in sig.choices.items()])

            row = [
                msg_name, msg_id_hex, id_type, cycle_time, sig_name,
                start_bit, length, byte_order, factor, offset, minimum, maximum,
                unit, mux_type, mux_val, val_desc
            ]
            ws.append(row)

    # Styling headers for readability
    for cell in ws[1]:
        cell.font = openpyxl.styles.Font(bold=True, color="FFFFFF")
        cell.fill = openpyxl.styles.PatternFill("solid", fgColor="4A90E2")

    try:
        os.makedirs(os.path.dirname(os.path.abspath(output_path)), exist_ok=True)
        wb.save(output_path)
        print(f"[OK] Generated {message_count} messages and {signal_count} signals.")
        print(f"\n============================================================")
        print(f"[DONE] Successfully generated Excel matrix:")
        print(f"Path: {output_path}")
        print(f"============================================================")
    except Exception as e:
        print(f"[ERROR] Error saving Excel file: {e}")
        sys.exit(1)

if __name__ == "__main__":
    if len(sys.argv) < 3:
        print("Usage: python dbc_to_excel.py <dbc_path> <output_excel_path>")
        sys.exit(1)
    dbc_to_excel(sys.argv[1], sys.argv[2])
